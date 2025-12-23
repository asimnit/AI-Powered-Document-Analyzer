"""
Excel Spreadsheet Parser

Extracts data from Excel spreadsheets (.xlsx, .xls) using openpyxl.
"""

import logging
from app.services.parsers.base import DocumentParser, ParserResult

logger = logging.getLogger(__name__)


class ExcelParser(DocumentParser):
    """
    Microsoft Excel parser
    
    Extracts text from spreadsheets:
    - All sheets
    - Cell values converted to text
    - Preserves sheet structure
    """
    
    def parse(self, file_path: str) -> ParserResult:
        """
        Extract text from Excel spreadsheet
        
        Args:
            file_path: Path to .xlsx or .xls file
            
        Returns:
            ParserResult with extracted text
        """
        if not self._validate_file(file_path):
            return ParserResult(
                text="",
                error=f"File not found or not readable: {file_path}"
            )
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path, data_only=True)
            text_parts = []
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
                
                # Extract data from rows
                for row in sheet.iter_rows(values_only=True):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        text_parts.append("\t".join(row_values))
            
            full_text = "\n".join(text_parts)
            
            return ParserResult(
                text=full_text,
                page_count=len(workbook.sheetnames),  # Each sheet is a "page"
                metadata={
                    "parser": "openpyxl",
                    "sheet_count": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames
                }
            )
            
        except Exception as e:
            error_msg = f"Failed to parse Excel spreadsheet: {str(e)}"
            logger.error(error_msg)
            return ParserResult(text="", error=error_msg)
    
    def supports(self, file_extension: str) -> bool:
        """Check if file extension is Excel"""
        return file_extension.lower() in [".xlsx", ".xls"]
